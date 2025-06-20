# Problem 3: Add average population per tract to countyData Dictionary


import openpyxl, pprint

print('Opening workbook...')
wb = openpyxl.load_workbook('censuspopdata.xlsx')
sheet = wb.active

countyData = {}

print('Reading rows...')
for row in range(2, sheet.max_row + 1):  # Skip header row
    state = sheet.cell(row=row, column=2).value
    county = sheet.cell(row=row, column=3).value
    pop = sheet.cell(row=row, column=4).value

    countyData.setdefault(state, {})
    countyData[state].setdefault(county, {'tracts': 0, 'pop': 0})

    countyData[state][county]['tracts'] += 1
    countyData[state][county]['pop'] += int(pop)


for state in countyData:
    for county in countyData[state]:
        data = countyData[state][county]
        data['avgPopPerTract'] = data['pop'] // data['tracts']


print('Saving results...')
with open('censusUpdated.py', 'w') as f:
    f.write('countyData = ' + pprint.pformat(countyData))

print('Done.')
